import json  # noqa: N999
from collections.abc import Iterable
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from utils import db


def _calculate_metrics(data: Iterable[dict]) -> dict:
    """Calculate metrics from the results data."""
    rejected = 0
    accepted = 0
    inconclusive = 0
    not_attempted = 0
    total = 0

    for row in data:
        if not isinstance(row, dict):
            continue
        total += 1
        status = row.get("qa_status")
        if status == "QaRejected":
            rejected += 1
        elif status == "QaAccepted":
            accepted += 1
        elif status == "QaInconclusive":
            inconclusive += 1
        elif status == "QaNotAttempted":
            not_attempted += 1

    return {
        "total": total,
        "rejected": rejected,
        "accepted": accepted,
        "inconclusive": inconclusive,
        "not_attempted": not_attempted,
    }


def _format_db_response(db_response: Iterable[tuple], experiment_type: str) -> list[dict]:
    """Format database response into a list of results."""
    if experiment_type == "dataset":
        return [value for row in db_response for value in json.loads(row[3]).values()]
    return [json.loads(row[3]) for row in db_response]


def _create_excel_export(df: pd.DataFrame) -> BytesIO | None:
    """Create an Excel export from the DataFrame with formatting."""
    try:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Results")
            ws = writer.sheets["Results"]

            if not df.empty and len(df.columns) > 0:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(fgColor="4472C4", fill_type="solid")

            for col in ws.columns:
                if col:
                    header_len = len(str(col[0].value or ""))
                    data_max = max((len(str(cell.value or "")) for cell in col[1:]), default=0)
                    max_len = max(header_len, data_max)

                    ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 50)
    except (ValueError, AttributeError, TypeError) as error:
        st.error(f"❌ Excel export failed: {error}")
        return None
    else:
        buffer.seek(0)
        return buffer


def _render_header(model: str, use_ocr: bool, experiment_type: str, ids: str) -> None:
    st.markdown(
        f"""
        Currently running an experiment with the following attributes:

        - Model: {model}
        - Use OCR: {bool(use_ocr)}
        - Type: {experiment_type}
        """
    )

    with st.expander("Open IDs"):
        st.text_area("IDs being processed: ", ids, disabled=True)

    st.markdown("<hr>", unsafe_allow_html=True)


def _add_pdf_links(df: pd.DataFrame) -> pd.DataFrame:
    if "file_reference" in df.columns:
        df = df.copy()
        df["View PDF"] = df["file_reference"].apply(
            lambda ref: f"/PDF_Viewer?reference_id={ref!s}" if pd.notna(ref) else None
        )
    return df


def _render_downloads(df: pd.DataFrame) -> None:
    btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 1])

    if btn_col1.button("Refresh Results", type="primary", width="stretch"):
        st.rerun()

    btn_col2.download_button(
        "📥 Download CSV",
        df.to_csv(index=False),
        "experiment_results.csv",
        "text/csv",
        width="stretch",
    )

    excel_buffer = _create_excel_export(df)
    if excel_buffer:
        btn_col3.download_button(
            "📥 Download XLSX",
            data=excel_buffer.getvalue(),
            file_name="experiment_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )


def _render_metrics(metrics: dict) -> None:
    col1, col2, col3, col4, col5 = st.columns(5)
    col2.metric("Accepted", metrics.get("accepted"))
    col1.metric("Rejected", metrics.get("rejected"))
    col3.metric("Not Attempted", metrics.get("not_attempted"))
    col4.metric("Inconclusive", metrics.get("inconclusive"))
    col5.metric("Total Processed", metrics.get("total"))


def _render_reset() -> None:
    st.markdown("<hr>", unsafe_allow_html=True)
    with st.popover("Reset experiment"):
        st.warning("This action cannot be undone.")
        confirm = st.checkbox("Yes, I understand the consequences")

        if st.button("Confirm delete", disabled=not confirm):
            st.success("All entries deleted.")
            db.reset_experiment()
            st.rerun()


def render_analytics_page() -> None:
    """Render the experiment analytics page."""
    st.title("Experiment Analytics")

    experiment = db.get_latest_experiment()
    if not experiment:
        st.info("No experiments found. Please run a new experiment in the 'Run' tab.")
        return

    experiment_id, experiment_type, ids, model, use_ocr, _override, _qalab_base_url, _timestamp = experiment
    _render_header(model=model, use_ocr=bool(use_ocr), experiment_type=experiment_type, ids=ids)

    data = db.get_results_by_experiment(experiment_id)
    qalab_results = _format_db_response(data, experiment_type=experiment_type)
    df = _add_pdf_links(pd.DataFrame(qalab_results))

    _render_downloads(df)
    _render_metrics(_calculate_metrics(qalab_results))

    st.dataframe(
        df,
        width="stretch",
        column_config={"View PDF": st.column_config.LinkColumn("View PDF", display_text="📄 Open")},
    )
    _render_reset()


render_analytics_page()
